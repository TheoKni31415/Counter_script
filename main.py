import os.path
import openpyxl
import datetime
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
print("Для продолжения работы введите пароль.")
password = input()

while password != "814039":
    print("Неверный пароль. Попробуйте еще раз.")
    password = input()

print("Введите артикул и количество единиц товаров по примеру '80000000 1'.")
print("После каждого ввода артикула и его количества нажимайте 'Enter'.")
print("Чтобы выгрузить итоговый список, после последнего введенного ариткула\nвведите слово 'список'.")
print()
print("Введите команду...")

command = input().lower()

dictart = {}

while command != "список":
    if len(command.split()) == 2 and (command.split()[0].isdigit() and command.split()[1].isdigit()) \
                                 and command.startswith("80") and len(command.split()[0]) == 8:
        dictart[command.split()[0]] = dictart.get(command.split()[0], 0) + int(command.split()[1])
        command = input()
    else:
        print("Ошибка некорректного ввода. Попробуйте еще раз.")
        command = input()

exellist = []
number = 1
for art in dictart:
    exellist.append(("", number, art, dictart[art]))
    number += 1

wb = openpyxl.Workbook()
list = wb.active
list.append(("", "", '', ''))
list.append(("", "", 'Артикул', 'Кол-во'))

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ft = Font(size=22)
al = Alignment(horizontal='center')

list['C2'].alignment = al
list['D2'].alignment = al
list['C2'].font = ft
list['D2'].font = ft

list.cell(row=2, column=3).border = thin_border
list.cell(row=2, column=4).border = thin_border

list.column_dimensions['A'].width = 5
list.column_dimensions['B'].width = 5
list.column_dimensions['C'].width = 20
list.column_dimensions['D'].width = 14
list.column_dimensions['G'].width = 22

for i in range (len(exellist)):
    list.append(exellist[i])
    for j in range (2, 5):
        list.cell(row=i + 3, column=j).border = thin_border
    list[f'B{i + 3}'].alignment = al
    list[f'C{i + 3}'].alignment = al
    list[f'D{i + 3}'].alignment = al
    list[f'B{i + 3}'].font = ft
    list[f'C{i + 3}'].font = ft
    list[f'D{i + 3}'].font = ft

this_day = datetime.datetime.today().strftime("%d.%m.%Y.")
count_file = 1
while os.path.exists(f"Архив/{this_day}.xlsx") == True:
    if this_day[-1] != ")":
        this_day += f" ({count_file})"
    else:
        count_file += 1
        this_day = f"{datetime.datetime.today().strftime('%d.%m.%Y.')} ({count_file})"

list['G2'].value = this_day
list['G2'].alignment = al
list['G2'].font = ft
list.cell(row=2, column=7).border = thin_border


count = 0
with open("Номер коробки.txt", "r") as file:
    file.seek(0)
    if len(file.read()) == 0:
        count = 1
    else:
        file.seek(0)
        count = int(file.read()) + 1
with open("Номер коробки.txt", "w") as file:
    file.write(str(count))

list['G3'].value = count
list['G3'].alignment = al
list['G3'].font = ft
list.cell(row=3, column=7).border = thin_border



wb.save(f"Архив/{this_day}.xlsx")

print("Список создан. Можете закрыть окно.")
s = input()







